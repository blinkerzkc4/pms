from openpyxl.utils import get_column_letter

from .models import Quantity, Unit


class QuantityExtractor:
    columns = [
        "s no",
        "description",
        "remarks",
        "length",
        "breadth",
        "height",
        "no",
        "unit",
        "quantity",
    ]

    def __init__(self, wb, sheet_name="", project=None):
        self.project = project
        if sheet_name:
            quantity_sheet = sheet_name
        else:
            quantity_sheet = [
                name for name in wb.sheetnames if "quantity" in name.lower()
            ][0]
        self.sheet = wb[quantity_sheet]

    def get_table_bound_header_map(self):
        for row_idx, row in enumerate(self.sheet.values, 1):
            row = [str(r) if r else "" for r in row]
            text = " ".join(r for r in row if r).lower()
            last_col = None

            if all(c in text for c in self.columns):
                last_col = [i + 1 for i, r in enumerate(row) if r.lower() == "remarks"][
                    0
                ]
                last_col = get_column_letter(last_col)
                column_map = {}
                for col in self.columns:
                    for i, val in enumerate(row):
                        if i in column_map.values():
                            continue
                        if val and col in str(val).lower():
                            column_map[col] = i
                            break
                break

        max_row = self.sheet.max_row
        return f"A{row_idx + 1}", f"{last_col}{max_row}", column_map

    def get_row_data(self, row_data, partid, subpartid, itemid, row_type):
        unit = row_data["unit"]
        if unit:
            unit, _ = Unit.objects.get_or_create(name=unit)
        data = {
            "type": row_type,
            "partid": partid,
            "subpartid": subpartid,
            "itemid": itemid,
            "s_no": row_data["s no"] or "",
            "description": row_data["description"] or "",
        }
        if row_type == "part":
            return data

        return {
            **data,
            "remarks": row_data["remarks"] or "",
            "no": row_data["no"] or "0.0",
            "length": row_data["length"] or "0.0",
            "breadth": row_data["breadth"] or "0.0",
            "height": row_data["height"] or "0.0",
            "area": row_data["area"] or "0.0",
            "quantity": row_data["quantity"] or "0.0",
            "unit": unit,
        }

    def extract(self):
        start, end, column_map = self.get_table_bound_header_map()
        partid, subpartid, itemid = 0, 0, 0
        all_data = []

        for row in self.sheet[start:end]:
            row_vals = [r.value for r in row]
            row_text = " ".join(str(r) for r in row_vals if r)
            row_data = {"area": 0}

            for header, i in column_map.items():
                row_data[header] = row_vals[i]

            if not row_text:
                continue
            desc = row_data["description"]
            sno = row_data["s no"]
            if desc and str(desc).lower().startswith("part "):
                partid += 1

                subpartid, itemid = 0, 0
                row_type = "PART"
            elif sno:
                subpartid += 1
                itemid = 0
                row_type = "SUB_PART"
            else:
                itemid += 1
                row_type = "ITEM"
            data = self.get_row_data(row_data, partid, subpartid, itemid, row_type)
            all_data.append(data)
        return all_data

    def clean(self, all_data):
        prev_unit = None
        for data in all_data[::-1]:
            part_id, subpart_id = data["partid"], data["subpartid"]
            if data["type"] == "ITEM":
                if data["unit"]:
                    prev_unit = data["unit"]
                    prev_partid, prev_subpartid = part_id, subpart_id
                elif (
                    prev_unit
                    and part_id == prev_partid
                    and subpart_id == prev_subpartid
                ):
                    data["unit"] = prev_unit
        return all_data

    def import_data(self):
        try:
            all_data = self.extract()
            all_data = self.clean(all_data)
            for data in all_data:
                try:
                    Quantity.objects.create(**data, project=self.project)
                except Exception as e:
                    print(e)

        except Exception as e:
            print(e)
