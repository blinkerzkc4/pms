import io
from datetime import datetime

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from project.models import Project
from utils.excel_generate import export_to_excel_old
from utils.number2words import number_in_words


class GenericExcelExporter:
    headers = []
    actual_heads = []
    title = "excel"

    def export(
        self, request, data, headers=[], actual_heads=[], title="excel", **kwargs
    ):
        headers = self.headers or headers
        actual_heads = self.actual_heads or actual_heads
        title = self.title or title
        context = {
            "headers": headers,
            "actual_heads": actual_heads,
            "title": title,
            "date": datetime.now().date(),
            "data": data,
            "company_name": request.user.assigned_municipality.office_name,
            "company_sub_name": request.user.assigned_municipality.sub_name,
            "company_address": request.user.assigned_municipality.office_address,
            "email": request.user.assigned_municipality.email,
            "phone": request.user.assigned_municipality.phone,
            "column_widths": [15, 40, 40, 40],
            "nepali": False,
        }
        response = export_to_excel_old(context)
        return response


class DistrictRateExcelExporter(GenericExcelExporter):
    headers = ["Item", "Unit", "Rate"]
    actual_heads = ["rate", "unit", "amount"]

    title = "district_rate"


class TransportExcelExporter(GenericExcelExporter):
    headers = [
        "Item",
        "Unit",
        "D1 Km",
        "D1 Rate",
        "D2 Km",
        "D2 Rate",
        "D3 Km",
        "D3 Rate ",
        "Total ",
        "Remark",
    ]
    actual_heads = [
        "rate",
        "unit",
        "road_1_amount",
        "road_1_distance",
        "road_2_amount",
        "road_2_distance",
        "road_3_amount",
        "road_3_distance",
        "transportation_rate",
    ]
    title = "district_rate"


class TrateExcelExporter(GenericExcelExporter):
    headers = [
        "Item",
        "Unit",
        "Base Rate",
        "Transportation",
        "Adopted Rate Inc. Transportation",
    ]
    actual_heads = [
        "rate",
        "unit",
        "amount",
        "transportation_rate",
        "total_rate",
    ]
    title = "district_tranport_rate"


class ResourceExcelExporter(GenericExcelExporter):
    headers = ["Item", "Unit", "Quantity"]
    actual_heads = ["title", "unit", "quantity"]

    title = "district_rate"


class QuantityExcelExporter:
    _sheet_name = "quantity_export.xlsx"
    headers = [
        "S No",
        "Description",
        "No",
        "Length",
        "Breadth",
        "Height",
        "quantity",
        "Unit",
        "remarks",
    ]

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    header_row_id = 9
    sheet_type = "quantities"

    def add_headers(self, sheet):
        for i, header in enumerate(self.headers, 1):
            cell = sheet.cell(self.header_row_id, i)
            cell.value = header
            cell.border = self.border
            cell.font = Font(bold=True)

    def add_rows(self, data, sheet):
        self.data_start_row_id = self.header_row_id + 1
        for i, row in enumerate(data, self.data_start_row_id):
            row_data = [
                row.get("s_no", ""),
                row.get("description", ""),
                row.get("no", ""),
                row.get("length", ""),
                row.get("breadth", ""),
                row.get("height", ""),
                row.get("quantity", ""),
                row["unit_value"]["name_eng"] if row["unit_value"] else "",
                row.get("remarks", ""),
            ]
            row_type = row.get("type")

            for j, d in enumerate(row_data, 1):
                cell = sheet.cell(i, j)
                if row_type == "PART":
                    # description is in the  2nd column, update if changed
                    cell.value = d if j == 2 else ""
                    cell.fill = PatternFill("solid", fgColor="00000000")
                    cell.font = Font(b=True, color="FFFFFFFF")
                else:
                    cell.value = d
                cell.border = self.border

    def _export(self, data):
        wb = Workbook()
        sheet = wb.active
        self.add_headers(sheet)
        self.add_rows(data, sheet)
        with io.BytesIO() as buffer:
            wb.save(buffer)
            content = buffer.getvalue()

        response = HttpResponse(
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=quantity.xlsx"
        return response

    def export(self, data):
        try:
            return self._export(data)

        except Exception as e:
            print(e)
            raise e


class BaseExcelExporter:
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_rowspan = 10
    max_length = 100
    min_lenth = 10

    def __init__(
        self,
        request,
        data,
        headers,
        title,
        header_row_span=1,
        sheet_name="download.xlsx",
        nepali_mode=False,
    ) -> None:
        self.request = request
        self.data = data
        self.title = title
        self.headers = headers
        self.header_row_span = header_row_span
        self.sheet_name = sheet_name
        self.nepali_mode = nepali_mode

    @property
    def total_columns(self):
        return len(self.columns_only)

    @property
    def header_row_id(self):
        return self.header_rowspan + 1

    @property
    def columns_only(self):
        columns = []
        for header in self.headers:
            if isinstance(header, dict):
                columns.extend(header.get("columns", []))
            else:
                columns.append(header)
        return columns

    @property
    def data_start_row_id(self):
        return self.header_row_id + self.header_row_span

    @property
    def context(self):
        return self.get_context()

    def get_context(self):
        return {
            "headers": self.headers,
            "title": self.title,
            "date": datetime.now().date(),
            "data": self.data,
            "project": Project.objects.filter(
                id=self.request.GET.get("project")
            ).first(),
            "company_name": self.request.user.assigned_municipality.office_name,
            "company_sub_name": self.request.user.assigned_municipality.sub_name,
            "company_address": self.request.user.assigned_municipality.office_address,
            "email": self.request.user.assigned_municipality.email,
            "phone": self.request.user.assigned_municipality.phone,
            "column_widths": [15, 40, 40, 40],
            "nepali": self.nepali_mode,
        }

    def add_headers(self, sheet):
        sheet[
            "A1"
        ].value = f'{self.context["company_name"]} \n {self.context["company_sub_name"]} \n {self.context["company_address"]} \n\n {self.context["title"]}'
        sheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=8,
            end_column=self.total_columns,
        )
        organization_font = Font(bold=True, size=18)
        organization_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        sheet["A1"].font = organization_font
        sheet["A1"].alignment = organization_alignment

        sheet["A9"].value = f"Project Title: {self.context['project'].name}"
        sheet.merge_cells(
            start_row=9,
            start_column=1,
            end_row=9,
            end_column=self.total_columns,
        )
        sheet["A9"].font = Font(bold=True)
        sheet["A9"].alignment = Alignment(horizontal="left")
        sheet["A10"].value = f"Location: {self.context['project'].municipality}"
        sheet.merge_cells(
            start_row=10,
            start_column=1,
            end_row=10,
            end_column=self.total_columns,
        )
        sheet["A10"].font = Font(bold=True)
        sheet["A10"].alignment = Alignment(horizontal="left")
        for i, width in enumerate(self.context["column_widths"], start=1):
            column_letter = chr(64 + i)
            sheet.column_dimensions[column_letter].width = width

    def add_row_headers(self, sheet):
        total_subheaders_added = 0
        for i, header in enumerate(self.headers, 1):
            if isinstance(header, dict):
                subheader_start_index = i
                subheader_columns = header.get("columns", [])
                for subheader_index, subheader in enumerate(
                    subheader_columns, subheader_start_index
                ):
                    cell = sheet.cell(self.header_row_id + 1, subheader_index)
                    cell.value = subheader
                    cell.border = self.border
                    cell.font = Font(bold=True)

                cell = sheet.cell(self.header_row_id, subheader_start_index)
                cell.value = header.get("title", "")
                cell.border = self.border
                cell.font = Font(bold=True)
                sheet.merge_cells(
                    start_row=self.header_row_id,
                    start_column=subheader_start_index,
                    end_row=self.header_row_id,
                    end_column=subheader_start_index + len(subheader_columns) - 1,
                )

                total_subheaders_added += len(subheader_columns)
            else:
                subheaders_index_to_add = total_subheaders_added - 1
                if subheaders_index_to_add < 0:
                    subheaders_index_to_add = 0
                column_id = i + subheaders_index_to_add
                cell = sheet.cell(self.header_row_id, column_id)
                cell.value = header
                cell.border = self.border
                cell.font = Font(bold=True)
                sheet.merge_cells(
                    start_row=self.header_row_id,
                    start_column=column_id,
                    end_row=self.header_row_id + self.header_row_span - 1,
                    end_column=column_id,
                )

    def add_rows(self, sheet):
        raise NotImplementedError

    def fix_columns_width(self, sheet: Worksheet):
        for col in sheet.columns:
            column_max_length = 0
            column = col[self.data_start_row_id].column_letter
            for cell in col:
                if cell.coordinate in sheet.merged_cells:
                    continue
                try:
                    if len(str(cell.value)) > column_max_length:
                        column_max_length = len(cell.value)
                except:
                    pass
            if column_max_length == 0:
                sheet.column_dimensions[column].width = (self.min_lenth + 2) * 1.2
            elif column_max_length < self.max_length:
                sheet.column_dimensions[column].width = (column_max_length + 2) * 1.2
            else:
                sheet.column_dimensions[column].width = (self.max_length + 2) * 1.2
                for cell in col:
                    old_alignment: Alignment = cell.alignment.copy()
                    cell.alignment = Alignment(
                        **{**old_alignment.__dict__, "wrapText": True}
                    )

    def export(self):
        wb = Workbook()
        sheet: Worksheet = wb.active
        self.add_headers(sheet)
        self.add_row_headers(sheet)
        self.add_rows(sheet)
        self.fix_columns_width(sheet)
        with io.BytesIO() as buffer:
            wb.save(buffer)
            content = buffer.getvalue()

        response = HttpResponse(
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f"attachment; filename={self.sheet_name}.xlsx"
        return response

    @classmethod
    def quick_export(cls, request, data, **kwargs):
        exporter = cls(request, data, **kwargs)
        return exporter.export()


class SORExcelExporter(BaseExcelExporter):
    def __init__(self, request, data, nepali_mode=False) -> None:
        super().__init__(
            request,
            data,
            headers=[
                "S.No.",
                "Specification No.",
                "Activity No.",
                "Description",
                "Unit",
                "Rate in NRs.",
            ],
            title="Summary of Rates",
            sheet_name="summary_of_rates_export",
            nepali_mode=nepali_mode,
        )

    def add_rows(self, sheet):
        for id, row in enumerate(self.data, self.data_start_row_id):
            row_type = row.get("type")
            row_data = [
                row.get("s_no_counter", ""),
                row.get("specification_no", ""),
                row.get("activity_no", ""),
                row.get("description", ""),
                row.get("unit", ""),
                row.get("amount", ""),
            ]
            if row_type == "TOPIC":
                cell = sheet.cell(id, 1)
                cell.value = row.get("description", "")
                cell.alignment = Alignment(horizontal="center")
                cell.border = self.border
                sheet.merge_cells(
                    start_row=id,
                    start_column=1,
                    end_row=id,
                    end_column=self.total_columns,
                )
            elif row_type == "SUBTOTAL":
                cell = sheet.cell(id, 1)
                cell.value = "Sub Total"
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="right")
                cell.border = self.border
                sheet.merge_cells(
                    start_row=id,
                    start_column=1,
                    end_row=id,
                    end_column=self.total_columns - 1,
                )
                cell = sheet.cell(id, self.total_columns)
                cell.value = row.get("amount", "")
                cell.font = Font(bold=True)
            else:
                for j, d in enumerate(row_data, 1):
                    cell = sheet.cell(id, j)
                    cell.border = self.border
                    cell.value = d


class AbstractOfCostExcelExporter(BaseExcelExporter):
    def __init__(self, request, data, nepali_mode=False) -> None:
        super().__init__(
            request,
            data,
            headers=[
                "S.No.",
                "Specification No.",
                "Activity No.",
                "Description",
                "Unit",
                "Rate in NRs.",
                "Percent in Total",
            ],
            title="Abstract of Cost",
            sheet_name="abstract_of_cost_export",
            nepali_mode=nepali_mode,
        )

    def add_rows(self, sheet):
        total = self.data.get("total")
        for id, row in enumerate(
            self.data.get("processed_summary_of_rates", []), self.data_start_row_id
        ):
            row_type = row.get("type")
            row_data = [
                row.get("s_no_counter", ""),
                row.get("specification_no", ""),
                row.get("activity_no", ""),
                row.get("description", ""),
                row.get("unit", ""),
                row.get("amount", ""),
                f"{(row.get('amount', 0) or 0) / (total or 1) * 100:.2f}%",
            ]
            if row_type == "TOPIC":
                cell = sheet.cell(id, 1)
                cell.value = row.get("description", "")
                cell.alignment = Alignment(horizontal="center")
                cell.border = self.border
                sheet.merge_cells(
                    start_row=id,
                    start_column=1,
                    end_row=id,
                    end_column=self.total_columns,
                )
            elif row_type == "SUBTOTAL":
                cell = sheet.cell(id, 1)
                cell.value = "Sub Total"
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="right")
                cell.border = self.border
                sheet.merge_cells(
                    start_row=id,
                    start_column=1,
                    end_row=id,
                    end_column=self.total_columns - 2,
                )
                cell = sheet.cell(id, self.total_columns - 1)
                cell.value = row.get("amount", "")
                cell.font = Font(bold=True)
                cell.border = self.border
                cell = sheet.cell(id, self.total_columns)
                cell.value = f'{(row.get("amount", 0) or 0) / (total or 1) * 100:.2f}%'
                cell.font = Font(bold=True)
                cell.border = self.border
            else:
                for j, d in enumerate(row_data, 1):
                    cell = sheet.cell(id, j)
                    cell.border = self.border
                    cell.value = d
        last_row = self.data_start_row_id + len(
            self.data.get("processed_summary_of_rates", [])
        )
        cell = sheet.cell(last_row, 1)
        cell.value = "Total"
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right")
        cell.border = self.border
        sheet.merge_cells(
            start_row=last_row,
            start_column=1,
            end_row=last_row,
            end_column=self.total_columns - 2,
        )
        cell = sheet.cell(last_row, self.total_columns - 1)
        cell.value = total
        cell.font = Font(bold=True)
        cell.border = self.border
        cell = sheet.cell(last_row, self.total_columns)
        cell.value = "100%"
        cell.font = Font(bold=True)
        cell.border = self.border


class SAOCExcelExporter(BaseExcelExporter):
    def __init__(self, request, data, nepali_mode=False) -> None:
        super().__init__(
            request,
            data,
            headers=["S.No.", "Description", "Amount", "In Words", "Remarks"],
            title="Summary of Abstract of Cost",
            sheet_name="summary_of_abstract_of_cost_export",
            nepali_mode=nepali_mode,
        )

    def add_rows(self, sheet):
        for id, row in enumerate(
            self.data.get("all_soac_data", []), self.data_start_row_id
        ):
            row_data = [
                f"{id - self.data_start_row_id + 1}",
                row.get("description", ""),
                row.get("amount", ""),
                number_in_words(row.get("amount", "") or 0),
                row.get("remarks", "") or "-",
            ]
            for j, d in enumerate(row_data, 1):
                cell = sheet.cell(id, j)
                cell.border = self.border
                cell.value = d

        last_row = self.data_start_row_id + len(self.data.get("all_soac_data", []))

        def add_other_data(title, amount, row_id):
            cell = sheet.cell(row_id, 1)
            cell.value = title
            cell.font = Font(bold=True)
            cell.border = self.border
            cell.alignment = Alignment(horizontal="right")
            sheet.merge_cells(
                start_row=row_id,
                start_column=1,
                end_row=row_id,
                end_column=self.total_columns - 3,
            )
            # amount
            cell = sheet.cell(row_id, self.total_columns - 2)
            cell.value = amount
            cell.font = Font(bold=True)
            cell.border = self.border
            # amount in words
            cell = sheet.cell(row_id, self.total_columns - 1)
            cell.value = number_in_words(amount)
            cell.border = self.border
            # remarks
            cell = sheet.cell(row_id, self.total_columns)
            cell.value = "-"
            cell.border = self.border

        add_other_data("Total", self.data.get("total"), last_row)
        add_other_data(
            "Provisional Sum", self.data.get("provisional_sum"), last_row + 1
        )
        add_other_data(
            "Total including P.S.", self.data.get("total_with_ps"), last_row + 1
        )
        add_other_data("Total VAT @ 13%", self.data.get("total_vat"), last_row + 1)
        add_other_data(
            "Grand Total with VAT", self.data.get("grand_total"), last_row + 1
        )


class QuantityEstimateExcelExporter(BaseExcelExporter):
    def __init__(self, request, data, nepali_mode=False) -> None:
        super().__init__(
            request,
            data,
            headers=[
                "S.No.",
                "Description of Works",
                "Unit",
                "No",
                {
                    "title": "Description of Quantity",
                    "columns": [
                        "Length",
                        "Breadth",
                        "Height",
                        "Quantity",
                    ],
                },
                "Type",
            ],
            header_row_span=2,
            title="Quantity Estimate",
            sheet_name="quantity_estimate_export",
            nepali_mode=nepali_mode,
        )

    def add_rows(self, sheet):
        for id, row in enumerate(self.data, self.data_start_row_id):
            row_data = [
                row.get("s_no", ""),
                row.get("description", ""),
                row["unit_value"]["name_eng"] if row["unit_value"] else "",
                row.get("no", ""),
                row.get("length", ""),
                row.get("breadth", ""),
                row.get("height", ""),
                row.get("quantity", ""),
                "D" if row.get("quantity_type", "") == "DIAMETER" else "",
            ]
            for j, d in enumerate(row_data, 1):
                cell = sheet.cell(id, j)
                cell.border = self.border
                cell.value = d
        last_row = self.data_start_row_id + len(self.data)
        cell = sheet.cell(last_row, 1)
        cell.value = "Total"
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right")
        cell.border = self.border
        sheet.merge_cells(
            start_row=last_row,
            start_column=1,
            end_row=last_row,
            end_column=self.total_columns - 2,
        )
        cell = sheet.cell(last_row, self.total_columns - 1)
        cell.value = sum([float(row.get("quantity", 0) or 0) for row in self.data])
        cell.font = Font(bold=True)
        cell.border = self.border
        cell = sheet.cell(last_row, self.total_columns)
        cell.value = "-"
        cell.font = Font(bold=True)
        cell.border = self.border


class CostEstimateExcelExporter(BaseExcelExporter):
    def __init__(self, request, data, nepali_mode=False) -> None:
        super().__init__(
            request,
            data,
            headers=[
                "S.No.",
                "Description of Works",
                "Unit",
                {
                    "title": "Description of Amount",
                    "columns": [
                        "Quantity",
                        "Rate",
                        "Amount",
                    ],
                },
            ],
            header_row_span=2,
            title="Cost Estimate",
            sheet_name="cost_estimate_export",
            nepali_mode=nepali_mode,
        )

    def add_rows(self, sheet):
        for id, row in enumerate(self.data.get("topics"), self.data_start_row_id):
            row_data = [
                row.get("s_no", ""),
                row.get("description", ""),
                row.get("unit", ""),
                row.get("quantity", ""),
                row.get("rate", ""),
                row.get("amount", ""),
            ]
            for j, d in enumerate(row_data, 1):
                cell = sheet.cell(id, j)
                cell.border = self.border
                cell.value = d

        last_row = self.data_start_row_id + len(self.data.get("topics", []))

        def add_other_data(title, amount, row_id):
            cell = sheet.cell(row_id, 1)
            cell.value = title
            cell.font = Font(bold=True)
            cell.border = self.border
            cell.alignment = Alignment(horizontal="right")
            sheet.merge_cells(
                start_row=row_id,
                start_column=1,
                end_row=row_id,
                end_column=3,
            )
            # amount
            cell = sheet.cell(row_id, 4)
            cell.value = amount
            cell.font = Font(bold=True)
            cell.border = self.border
            sheet.merge_cells(
                start_row=row_id,
                start_column=4,
                end_row=row_id,
                end_column=6,
            )

        add_other_data("Overhead 15%", self.data.get("overhead"), last_row)
        last_row += 1
        add_other_data("Subtotal", self.data.get("subtotal"), last_row)
        last_row += 1
        add_other_data("Add 1.5% Contingency", self.data.get("contingency"), last_row)
        last_row += 1
        add_other_data("VAT", self.data.get("vat"), last_row)
        last_row += 1
        add_other_data("Grand Total", self.data.get("grand_total"), last_row)
        last_row += 1

        cell = sheet.cell(last_row, 1)
        cell.value = "In Words"
        cell.font = Font(bold=True)
        cell.border = self.border
        cell.alignment = Alignment(horizontal="right")
        # amount
        cell = sheet.cell(last_row, 2)
        cell.value = number_in_words(self.data.get("grand_total", 0))
        cell.border = self.border
        sheet.merge_cells(
            start_row=last_row,
            start_column=2,
            end_row=last_row,
            end_column=6,
        )
