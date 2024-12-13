"""
-- Created by Bikash Saud
-- Created on 2023-06-28
"""

import io
import json
from decimal import Decimal

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from utils.file_utils import get_media_file_path
from utils.nepali_nums import nepali_nums


def get_excel_response(data):
    workbook = Workbook()
    sheet: Worksheet = workbook.active
    sheet[
        "A1"
    ].value = f'{data["company_name"]} \n {data["company_sub_name"]} \n {data["company_address"]} \n\n {data["title"]}'
    cell = sheet.cell(row=8, column=len(data["display_fields"]) + 1)
    sheet.merge_cells(f"A1:{cell.column_letter}{cell.row}")
    organization_font = Font(bold=True, size=18)
    organization_alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    sheet["A1"].font = organization_font
    sheet["A1"].alignment = organization_alignment

    for i, width in enumerate(data["column_widths"], start=1):
        column_letter = chr(64 + i)
        sheet.column_dimensions[column_letter].width = width

    header_font = Font(bold=True)

    cell = sheet.cell(row=9, column=1)
    cell.value = "क्र. स." if data.get("nepali", True) else "S No."
    cell.font = header_font

    for col, header in enumerate(
        [display_field[1] for display_field in data["display_fields"]], start=2
    ):
        cell = sheet.cell(row=9, column=col)
        cell.value = header
        cell.font = header_font

    counter = 0
    data_fields = [display_field[0] for display_field in data["display_fields"]]
    get_data = data["data"]
    export_data = get_data(data["request"]) if callable(get_data) else get_data
    print(export_data)
    for row_index, data_row_val in enumerate(export_data, start=10):
        counter += 1
        row_data = [nepali_nums(counter) if data.get("nepali", True) else counter]
        for i in data_fields:
            print(i)
            if isinstance(i, str):
                val = data_row_val[i]
            else:
                val = i(data_row_val)
            print(val)
            if val is None:
                val = ""
            elif isinstance(val, Decimal):
                val = round(val, 3)
            else:
                val = str(val)
            row_data.append(val)

        for col_index, value in enumerate(row_data, start=1):
            sheet.cell(row=row_index, column=col_index).value = value

    for col in sheet.columns:
        max_length = 0
        column = col[8].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[column].width = adjusted_width

    with io.BytesIO() as buffer:
        workbook.save(buffer)
        content = buffer.getvalue()

    response = HttpResponse(
        content=content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=" + f"{data['title']}.xlsx"
    return response


def export_to_excel_old(data):
    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet[
            "A1"
        ].value = f'{data["company_name"]} \n {data["company_sub_name"]} \n {data["company_address"]} \n\n {data["title"]}'
        sheet.merge_cells("A1:D8")
        organization_font = Font(bold=True, size=18)
        organization_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        sheet["A1"].font = organization_font
        sheet["A1"].alignment = organization_alignment
        for i, width in enumerate(data["column_widths"], start=1):
            column_letter = chr(64 + i)
            sheet.column_dimensions[column_letter].width = width

        header_font = Font(bold=True)
        for col, header in enumerate(data["headers"], start=1):
            cell = sheet.cell(row=9, column=col + 1)
            cell.value = header
            cell.font = header_font
        cell = sheet.cell(row=9, column=1)
        cell.value = "क्र. स." if data.get("nepali", True) else "S No."
        cell.font = header_font

        counter = 0
        for row_index, data_row_val in enumerate(data["data"], start=10):
            counter += 1
            row_data = [nepali_nums(counter) if data.get("nepali", True) else counter]
            for i in data["actual_heads"]:
                if isinstance(data_row_val, dict):
                    val = data_row_val[i]
                else:
                    val = getattr(data_row_val, i)
                if val is None:
                    val = ""
                elif isinstance(val, Decimal):
                    val = round(val, 3)
                else:
                    val = str(val)
                row_data.append(val)
            for col_index, value in enumerate(row_data, start=1):
                sheet.cell(row=row_index, column=col_index).value = value

        with io.BytesIO() as buffer:
            workbook.save(buffer)
            content = buffer.getvalue()
        response = HttpResponse(
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f"attachment; filename={data['title']}.xlsx"
        return response
    except Exception as e:
        # print(f"Error Generating excel sheet: {e}")
        raise e
