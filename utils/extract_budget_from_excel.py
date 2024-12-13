import json
import locale

import openpyxl

QUARTERS_TOPICS_DICT = {
    "सि.नं.": "sn_no",
    "कार्यक्रम/आयोजना/क्रियाकलापको नाम": "activity_name",
    "उप क्षेत्र": "sub_topic",
    "खर्च शीर्षक": "expense_title",
    "स्रोत": "source",
    "लक्ष": "target",
    "इकाई": "unit",
    "विनियोजन": {
        "पहिलो त्रैमासिक": "first_quarter",
        "दोस्रो त्रैमासिक": "second_quarter",
        "तेस्रो त्रैमासिक": "third_quarter",
        "चौथो त्रैमासिक": "fourth_quarter",
        "जम्मा": "total",
    },
}

STARTING_COL_NAMES = ["सि.नं.", "सिनं", "सि.नं", "सिनं.", "S.N.", "S.N", "SN", "SN."]


def extract_data_from_four_quarters_excel(excel_file, amount_is_thousands=False):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    sub_topics_dict = {}
    sub_topics_row, topics_row = None, None
    found_sn = False

    # Search for the "सि.नं." marker in the first two columns
    for row in sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=2):
        for cell in row:
            if cell.value in STARTING_COL_NAMES:
                found_sn = True
                topics_row = row
                break
        if found_sn:
            break
    if not found_sn:
        raise Exception("Could not find the सि.नं. marker in the first two columns")
    else:
        topics_row_number = topics_row[0].row
        start_index_row = topics_row_number + 2
        start_rows = list(sheet.iter_rows(min_row=start_index_row))
        rows_to_process = start_rows[:-1]
        budget = {}  # Main dictionary
        current_budget_head = None

        for row in rows_to_process:
            if all(cell.value is None for cell in row[4:7]):
                for bheads in row:
                    if bheads.value is not None:
                        current_budget_head = bheads.value
                        budget[current_budget_head] = []
            else:
                row_data = {}
                for index, key in enumerate(QUARTERS_TOPICS_DICT.keys()):
                    if key != "विनियोजन":
                        value = row[index].value
                        row_data[QUARTERS_TOPICS_DICT[key]] = value
                    else:
                        row_data["allocation"] = {}
                        for sub_key_index, sub_key in enumerate(
                            QUARTERS_TOPICS_DICT[key].keys()
                        ):
                            value = (row[index + sub_key_index].value) or 0
                            try:
                                value = int(value)
                            except ValueError:
                                value = value.replace(" ", "")
                                locale.setlocale(locale.LC_ALL, "en_US.UTF-8")
                                value = locale.atof(value)
                            row_data["allocation"][
                                QUARTERS_TOPICS_DICT[key][sub_key]
                            ] = (value * 1000 if amount_is_thousands else value)

                if current_budget_head is not None:
                    budget[current_budget_head].append(row_data)
    return budget


def main():
    budget = extract_data_from_four_quarters_excel("./test_data/jsmc import.xlsx", True)

    with open("./test_data/budget.json", "w", encoding="utf-8") as json_file:
        json.dump(budget, json_file, ensure_ascii=False, indent=4)

    print(json.dumps(list(budget.items())[-1], indent=4, ensure_ascii=False))


if __name__ == "__main__":
    main()
