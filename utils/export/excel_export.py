import io
from datetime import datetime

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet

from project.models import Project


class BaseExcelExporter:
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_rowspan = 10
    max_length = 100
    min_lenth = 10

    def __init__(
        self,
        request,
        data,
        headers,
        title,
        header_row_span=1,
        sheet_name="download.xlsx",
        nepali_mode=False,
    ) -> None:
        self.request = request
        self.data = data
        self.title = title
        self.headers = headers
        self.header_row_span = header_row_span
        self.sheet_name = sheet_name
        self.nepali_mode = nepali_mode

    @property
    def total_columns(self):
        return len(self.columns_only)

    @property
    def header_row_id(self):
        return self.header_rowspan + 1

    @property
    def columns_only(self):
        columns = []
        for header in self.headers:
            if isinstance(header, dict):
                columns.extend(header.get("columns", []))
            else:
                columns.append(header)
        return columns

    @property
    def data_start_row_id(self):
        return self.header_row_id + self.header_row_span

    @property
    def context(self):
        return self.get_context()

    def get_context(self):
        return {
            "headers": self.headers,
            "title": self.title,
            "date": datetime.now().date(),
            "data": self.data,
            "project": Project.objects.filter(
                id=self.request.GET.get("project")
            ).first(),
            "company_name": self.request.user.assigned_municipality.office_name,
            "company_sub_name": self.request.user.assigned_municipality.sub_name,
            "company_address": self.request.user.assigned_municipality.office_address,
            "email": self.request.user.assigned_municipality.email,
            "phone": self.request.user.assigned_municipality.phone,
            "column_widths": [15, 40, 40, 40],
            "nepali": self.nepali_mode,
        }

    def add_headers(self, sheet):
        sheet["A1"].value = (
            f'{self.context["company_name"]} \n {self.context["company_sub_name"]} \n {self.context["company_address"]} \n\n {self.context["title"]}'
        )
        organization_font = Font(bold=True, size=18)
        organization_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        sheet["A1"].font = organization_font
        sheet["A1"].alignment = organization_alignment

        if self.context["project"]:
            sheet["A9"].value = f"Project Title: {self.context['project'].name}"
            sheet["A9"].font = Font(bold=True)
            sheet["A9"].alignment = Alignment(horizontal="left")
            sheet["A10"].value = f"Location: {self.context['project'].municipality}"
            sheet["A10"].font = Font(bold=True)
            sheet["A10"].alignment = Alignment(horizontal="left")
            sheet.merge_cells(
                start_row=1,
                start_column=1,
                end_row=8,
                end_column=self.total_columns,
            )
            sheet.merge_cells(
                start_row=9,
                start_column=1,
                end_row=9,
                end_column=self.total_columns,
            )
            sheet.merge_cells(
                start_row=10,
                start_column=1,
                end_row=10,
                end_column=self.total_columns,
            )
        else:
            sheet.merge_cells(
                start_row=1,
                start_column=1,
                end_row=10,
                end_column=self.total_columns,
            )
        for i, width in enumerate(self.context["column_widths"], start=1):
            column_letter = chr(64 + i)
            sheet.column_dimensions[column_letter].width = width

    def add_row_headers(self, sheet):
        total_subheaders_added = 0
        for i, header in enumerate(self.headers, 1):
            if isinstance(header, dict):
                subheader_start_index = i
                subheader_columns = header.get("columns", [])
                for subheader_index, subheader in enumerate(
                    subheader_columns, subheader_start_index
                ):
                    cell = sheet.cell(self.header_row_id + 1, subheader_index)
                    cell.value = subheader
                    cell.border = self.border
                    cell.font = Font(bold=True)

                cell = sheet.cell(self.header_row_id, subheader_start_index)
                cell.value = header.get("title", "")
                cell.border = self.border
                cell.font = Font(bold=True)
                sheet.merge_cells(
                    start_row=self.header_row_id,
                    start_column=subheader_start_index,
                    end_row=self.header_row_id,
                    end_column=subheader_start_index + len(subheader_columns) - 1,
                )

                total_subheaders_added += len(subheader_columns)
            else:
                subheaders_index_to_add = total_subheaders_added - 1
                if subheaders_index_to_add < 0:
                    subheaders_index_to_add = 0
                column_id = i + subheaders_index_to_add
                cell = sheet.cell(self.header_row_id, column_id)
                cell.value = header
                cell.border = self.border
                cell.font = Font(bold=True)
                sheet.merge_cells(
                    start_row=self.header_row_id,
                    start_column=column_id,
                    end_row=self.header_row_id + self.header_row_span - 1,
                    end_column=column_id,
                )

    def add_rows(self, sheet):
        raise NotImplementedError

    def fix_columns_width(self, sheet: Worksheet):
        for col in sheet.columns:
            column_max_length = 0
            try:
                column = col[self.header_row_id].column_letter
            except:
                continue
            for cell in col:
                if cell.coordinate in sheet.merged_cells:
                    continue
                try:
                    if len(str(cell.value)) > column_max_length:
                        column_max_length = len(cell.value)
                except:
                    pass
            if column_max_length == 0:
                sheet.column_dimensions[column].width = (self.min_lenth + 2) * 1.2
            elif column_max_length < self.max_length:
                sheet.column_dimensions[column].width = (column_max_length + 2) * 1.2
            else:
                sheet.column_dimensions[column].width = (self.max_length + 2) * 1.2
                for cell in col:
                    old_alignment: Alignment = cell.alignment.copy()
                    cell.alignment = Alignment(
                        **{**old_alignment.__dict__, "wrapText": True}
                    )

    def export(self):
        wb = Workbook()
        sheet: Worksheet = wb.active
        self.add_headers(sheet)
        self.add_row_headers(sheet)
        self.add_rows(sheet)
        self.fix_columns_width(sheet)
        with io.BytesIO() as buffer:
            wb.save(buffer)
            content = buffer.getvalue()

        response = HttpResponse(
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f"attachment; filename={self.sheet_name}.xlsx"
        return response

    @classmethod
    def quick_export(cls, request, data, **kwargs):
        exporter = cls(request, data, **kwargs)
        return exporter.export()


class NormExcelExporter(BaseExcelExporter):
    def __init__(self, request, data, nepali_mode=False) -> None:
        super().__init__(
            request,
            data,
            headers=[
                "S.No.",
                "Activity Number",
                "Specification Number",
                "Description",
                "Unit",
                "Rate per Unit",
                "Remarks",
            ],
            title="Norms List",
            sheet_name="norms_list_export",
            nepali_mode=nepali_mode,
        )

    def add_rows(self, sheet):
        for i, row in enumerate(self.data):
            sheet.cell(self.data_start_row_id + i, 1).value = i + 1
            sheet.cell(self.data_start_row_id + i, 2).value = row.get("activity_no")
            sheet.cell(self.data_start_row_id + i, 3).value = row.get(
                "specification_no"
            )
            sheet.cell(self.data_start_row_id + i, 4).value = row.get(
                "description"
            ) or row.get("description_eng")
            sheet.cell(self.data_start_row_id + i, 5).value = row.get(
                "unit_name"
            ) or row.get("unit_name_eng")
            sheet.cell(self.data_start_row_id + i, 6).value = float(
                f"{row.get('analysed_rate'):.2f}"
            )
            sheet.cell(self.data_start_row_id + i, 7).value = row.get("remarks")

            for j in range(1, self.total_columns + 1):
                sheet.cell(self.data_start_row_id + i, j).border = self.border
