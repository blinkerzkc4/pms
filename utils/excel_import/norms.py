import logging

from openpyxl import load_workbook

from utils.excel_import.excel_import_utils import get_no_children_snos

logger = logging.getLogger("Norms Importer")
logging.basicConfig(level=logging.INFO)


def extract_data(file, worksheet_name=None):
    workbook = load_workbook(file)
    sheet = workbook[worksheet_name] if worksheet_name else workbook.active

    data = {}

    initial_data = []

    for row in sheet.iter_rows(values_only=True, min_row=4):
        (
            s_no,
            activity_no,
            spec_ci_no,
            description,
            description_eng,
            unit_value,
            unit,
            work_force_type,
            work_force_unit,
            work_force_rate,
            work_force_quantity,
            work_force_amount,
            materials_type,
            materials_unit,
            materials_rate,
            materials_quantity,
            materials_amount,
            equipment_type,
            equipment_unit,
            equipment_rate,
            equipment_quantity,
            equipment_amount,
            remarks,
        ) = row

        if activity_no and spec_ci_no:
            norm_data = {
                "item_id": s_no,
                "activity_no": activity_no,
                "specification_no": spec_ci_no,
                "description": description,
                "description_eng": description_eng,
                "unit_value": unit_value or 1,
                "unit": unit,
                "remarks": remarks,
                "active": True,
            }

            if work_force_type:
                norm_data["labour_component"] = [
                    {
                        "category_name": work_force_type,
                        "component_type": "LABOUR",
                        "rate": work_force_rate,
                        "quantity": work_force_quantity,
                        "unit": work_force_unit,
                        "amount": work_force_amount,
                    }
                ]

            if materials_type:
                norm_data["material_component"] = [
                    {
                        "category_name": materials_type,
                        "component_type": "MATERIAL",
                        "rate": materials_rate,
                        "quantity": materials_quantity,
                        "unit": materials_unit,
                        "amount": materials_amount,
                    }
                ]

            if equipment_type:
                norm_data["equipment_component"] = [
                    {
                        "category_name": equipment_type,
                        "component_type": "EQUIPMENT",
                        "rate": equipment_rate,
                        "quantity": equipment_quantity,
                        "unit": equipment_unit,
                        "amount": equipment_amount,
                    }
                ]
            initial_data.append(norm_data)
        elif len(initial_data) > 0:
            if work_force_type:
                initial_data[-1]["labour_component"].append(
                    {
                        "category_name": work_force_type,
                        "component_type": "LABOUR",
                        "rate": work_force_rate,
                        "quantity": work_force_quantity,
                        "unit": work_force_unit,
                        "amount": work_force_amount,
                    }
                )

            if materials_type:
                initial_data[-1]["material_component"].append(
                    {
                        "category_name": materials_type,
                        "component_type": "MATERIAL",
                        "rate": materials_rate,
                        "quantity": materials_quantity,
                        "unit": materials_unit,
                        "amount": materials_amount,
                    }
                )

            if equipment_type:
                initial_data[-1]["equipment_component"].append(
                    {
                        "category_name": equipment_type,
                        "component_type": "EQUIPMENT",
                        "rate": equipment_rate,
                        "quantity": equipment_quantity,
                        "unit": equipment_unit,
                        "amount": equipment_amount,
                    }
                )

    data = {str(norms_data["item_id"]): norms_data for norms_data in initial_data}

    return data


def process_data(data):
    processed_data = []

    no_children_snos = get_no_children_snos(data.keys())

    for no_children_sno in no_children_snos:
        norm_data = {}
        norm_data.update(data[no_children_sno])

        level_codes = no_children_sno.split(".")

        if len(level_codes) == 1:
            norm_data["item_id"] = norm_data["item_id"]
            norm_data["activity_no"] = norm_data["activity_no"]
            norm_data["specification_no"] = norm_data["specification_no"]
            norm_data["description"] = norm_data["description"]
            norm_data["description_eng"] = norm_data["description_eng"]

            norm_data["sub_part_id"] = ""
            norm_data["subpart_activity_no"] = ""
            norm_data["subpart_specification_no"] = ""
            norm_data["subpart_description"] = ""
            norm_data["subpart_description_eng"] = ""

            norm_data["part_id"] = norm_data["item_id"]
            norm_data["part_activity_no"] = norm_data["activity_no"]
            norm_data["part_specification_no"] = norm_data["specification_no"]
            norm_data["part_description"] = norm_data["description"]
            norm_data["part_description_eng"] = norm_data["description_eng"]

        elif len(level_codes) == 2:
            norm_data["activity_no"] = norm_data["activity_no"]
            norm_data["specification_no"] = norm_data["specification_no"]

            norm_data["sub_part_id"] = norm_data["item_id"]
            norm_data["subpart_activity_no"] = norm_data["activity_no"]
            norm_data["subpart_specification_no"] = norm_data["specification_no"]
            norm_data["subpart_description"] = norm_data["description"]
            norm_data["subpart_description_eng"] = norm_data["description_eng"]

            norm_data["item_id"] = data[level_codes[0]]["item_id"]
            norm_data["description"] = data[level_codes[0]]["description"]
            norm_data["description_eng"] = data[level_codes[0]]["description_eng"]
            norm_data["part_id"] = data[level_codes[0]]["item_id"]
            norm_data["part_activity_no"] = data[level_codes[0]]["activity_no"]
            norm_data["part_specification_no"] = data[level_codes[0]][
                "specification_no"
            ]
            norm_data["part_description"] = data[level_codes[0]]["description"]
            norm_data["part_description_eng"] = data[level_codes[0]]["description_eng"]
        elif len(level_codes) == 3:
            norm_data["activity_no"] = norm_data["activity_no"]
            norm_data["specification_no"] = norm_data["specification_no"]
            norm_data["item_description"] = norm_data["description"]
            norm_data["item_description_eng"] = norm_data["description_eng"]

            norm_data["sub_part_id"] = data[f"{level_codes[0]}.{level_codes[1]}"][
                "item_id"
            ]
            norm_data["subpart_activity_no"] = data[
                f"{level_codes[0]}.{level_codes[1]}"
            ]["activity_no"]
            norm_data["subpart_specification_no"] = data[
                f"{level_codes[0]}.{level_codes[1]}"
            ]["specification_no"]
            norm_data["subpart_description"] = data[
                f"{level_codes[0]}.{level_codes[1] }"
            ]["description"]
            norm_data["subpart_description_eng"] = data[
                f"{level_codes[0]}.{level_codes[1]}"
            ]["description_eng"]

            norm_data["item_id"] = data[level_codes[0]]["item_id"]
            norm_data["description"] = data[level_codes[0]]["description"]
            norm_data["description_eng"] = data[level_codes[0]]["description_eng"]
            norm_data["part_id"] = data[level_codes[0]]["item_id"]
            norm_data["part_activity_no"] = data[level_codes[0]]["activity_no"]
            norm_data["part_specification_no"] = data[level_codes[0]][
                "specification_no"
            ]
            norm_data["part_description"] = data[level_codes[0]]["description"]
            norm_data["part_description_eng"] = data[level_codes[0]]["description_eng"]

        json_data = {
            "activity": "",
            "activity_no": "",
            "item_id": "",
            "description_unicode": "",
            "unit": "",
            "unit_value": "",
            "description": "",
            "subpart_description": "",
            "item_description": "",
            "part_id": "",
            "sub_part_id": "",
            "specification_no": "",
            "part_specification_no": "",
            "subpart_specification_no": "",
            "item_specification_no": "",
            "remarks": "",
            "labour_component": [],
            "material_component": [],
            "equipment_component": [],
            "active": True,
        }

        for json_key, json_default_value in json_data.items():
            norm_data.setdefault(json_key, json_default_value)

        processed_data.append(norm_data)

    return processed_data


def import_norms_data(file, worksheet_name):
    logger.info(f"Importing norms data from {file}")
    data = extract_data(file, worksheet_name)

    return process_data(data)


if __name__ == "__main__":
    data = import_norms_data("test_data/Import Excel Sheets.xlsx", "Norms")
    import json

    print(json.dumps(data, indent=4))
