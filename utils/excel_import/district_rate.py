import logging

from openpyxl import load_workbook

from utils.excel_import.excel_import_utils import get_no_children_snos

logger = logging.getLogger("District Rate Importer")
logging.basicConfig(level=logging.INFO)


def extract_data(file, worksheet_name=None):
    workbook = load_workbook(file)
    sheet = workbook[worksheet_name] if worksheet_name else workbook.active

    data = {}

    for row in sheet.iter_rows(values_only=True, min_row=3):
        (
            s_no,
            category,
            content,
            content_unicode,
            unit,
            source,
            area,
            rate_1,
            rate_2,
            rate_3,
        ) = row
        data[str(s_no)] = {
            "s_no": s_no,
            "category": category,
            "content": content,
            "content_unicode": content_unicode,
            "unit": unit,
            "source": source,
            "rate_area": area,
            "rate_1": rate_1,
            "rate_2": rate_2,
            "rate_3": rate_3,
        }

    return data


def process_data(data):
    processed_data = []

    INPUT_TYPES = ["TOPIC", "SUBTOPIC", "ITEM"]

    no_children_snos = get_no_children_snos(data.keys())

    for no_parent_s_no in no_children_snos:
        dr_data = {}
        dr_data.update(data[no_parent_s_no])

        level_codes = no_parent_s_no.split(".")
        dr_data["input_type"] = INPUT_TYPES[len(level_codes) - 1]

        dr_data["title"] = dr_data["content"]
        dr_data["title_eng"] = dr_data["content_unicode"]
        dr_data["input"] = dr_data["content"]
        dr_data["input_unicode"] = dr_data["content_unicode"]

        print(data)
        print(level_codes)
        if len(level_codes) == 1 and level_codes[-1] != "0":
            dr_data["item"] = ""
            dr_data["item_unicode"] = ""
            dr_data["sub_topic"] = ""
            dr_data["sub_topic_unicode"] = ""
            dr_data["topic"] = dr_data["content"]
            dr_data["topic_unicode"] = dr_data["content_unicode"]
        elif len(level_codes) == 2 and level_codes[-1] != "0":
            dr_data["item"] = ""
            dr_data["item_unicode"] = ""
            dr_data["sub_topic"] = dr_data["content"]
            dr_data["sub_topic_unicode"] = dr_data["content_unicode"]
            try:
                dr_data["topic"] = data[level_codes[0]]["content"]
                dr_data["topic_unicode"] = data[level_codes[0]]["content_unicode"]
            except KeyError:
                dr_data["topic"] = data[f"{level_codes[0]}.0"]["content"]
                dr_data["topic_unicode"] = data[f"{level_codes[0]}.0"][
                    "content_unicode"
                ]
        elif len(level_codes) == 3 and level_codes[-1] != "0":
            dr_data["item"] = dr_data["content"]
            dr_data["item_unicode"] = dr_data["content_unicode"]
            dr_data["sub_topic"] = data[f"{level_codes[0]}.{level_codes[1]}"]["content"]
            dr_data["sub_topic_unicode"] = data[f"{level_codes[0]}.{level_codes[1]}"][
                "content_unicode"
            ]
            try:
                dr_data["topic"] = data[level_codes[0]]["content"]
                dr_data["topic_unicode"] = data[level_codes[0]]["content_unicode"]
            except KeyError:
                dr_data["topic"] = data[f"{level_codes[0]}.0"]["content"]
                dr_data["topic_unicode"] = data[f"{level_codes[0]}.0"][
                    "content_unicode"
                ]

        json_data = {
            "category": "",
            "title": "",
            "title_eng": "",
            "description": "",
            "amount": "",
            "topic": "default",
            "topic_unicode": "default",
            "sub_topic": "",
            "sub_topic_unicode": "",
            "item": "",
            "item_unicode": "",
            "rate_1": "",
            "rate_2": "",
            "rate_3": "",
            "unit": "",
            "input": "",
            "input_unicode": "",
            "source": "",
            "rate_area": "",
        }

        for key in json_data.keys():
            if key not in dr_data.keys():
                logger.warning(f"Key {key} not found in dr_data. Setting to empty.")
                dr_data[key] = json_data[key]

        processed_data.append(dr_data)

    return processed_data


def import_district_rate_data(file, worksheet_name=None):
    logger.info(f"Importing data from {file}")
    data = extract_data(file, worksheet_name)
    logger.info("Data extracted. Processing data.")
    return process_data(data)


if __name__ == "__main__":
    data = import_district_rate_data(
        "test_data/Import Excel Sheets.xlsx", "District Rate"
    )
    import json

    print(json.dumps(data, indent=4))
