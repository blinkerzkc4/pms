import json
import re

import openpyxl

from utils.excel_import.constants import QuantityEstimateImportConstants
from utils.excel_import.excel_import_utils import get_columns_list, get_total_row_span
from utils.excel_import.exceptions import SNoNotFoundException
from utils.nepali_nums import english_nums


def extract_data_from_new_template(excel_file):
    pattern = r"^\((\d+(\.\d+)?)\)$"
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = workbook.worksheets[0]
    sub_topics_dict = {}
    sub_topics_row, topics_row = None, None

    # Search for the "सि.नं." marker in the first two columns
    for row in sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=2):
        for cell in row:
            if cell.value in QuantityEstimateImportConstants.STARTING_COL_NAMES:
                topics_row = row
                break
        if topics_row:
            break
    if not topics_row:
        raise SNoNotFoundException(
            "Could not find the सि.नं. marker in the first two columns.\nMarkers: {}".format(
                QuantityEstimateImportConstants.STARTING_COL_NAMES
            )
        )

    start_index_row = topics_row[-1].row + get_total_row_span(
        QuantityEstimateImportConstants.QUANTITY_ESTIMATE_TOPICS_DICT
    )

    quantity_estimate_data = []

    rows_to_process = list(sheet.iter_rows(min_row=start_index_row))

    columns_list = get_columns_list(
        QuantityEstimateImportConstants.QUANTITY_ESTIMATE_TOPICS_DICT
    )

    for row in rows_to_process:
        row_data = {"type": "data"}
        for index, dict_items in enumerate(
            QuantityEstimateImportConstants.QUANTITY_ESTIMATE_TOPICS_DICT.items()
        ):
            dict_key, dict_value = dict_items
            if not isinstance(dict_value, dict):
                column_number = columns_list.index(dict_value)
                row_value = row[column_number].value
                row_data[dict_value] = english_nums(row_value)
                # row_data[dict_value] = row_value
            else:
                row_data[dict_value["code"]] = {}
                for sub_key_index, sub_key_items in enumerate(
                    dict_value["fields"].items()
                ):
                    sub_key, sub_value = sub_key_items
                    column_number = columns_list.index(sub_value)
                    value = row[column_number].value
                    if value is not None:
                        try:
                            float(value)
                        except:
                            match = re.match(pattern, str(value))
                            if match:
                                value = float(match.group(1))
                            else:
                                value = 0
                        value = float(f"{value:.5f}")
                    row_data[dict_value["code"]][sub_value] = value

        if (row[0].value is None or str(row[0].value).strip() == "") and len(
            quantity_estimate_data
        ) > 0:
            last_s_no = quantity_estimate_data[-1]["s_no"]
            if len(last_s_no.split(".")) > 1:
                last_s_no = last_s_no.split(".")
                last_s_no[-1] = str(int(last_s_no[-1]) + 1)
                last_s_no = ".".join(last_s_no)
            else:
                last_s_no = f"{last_s_no}.1"
            row_data["s_no"] = last_s_no

        quantity_estimate_data.append(row_data)

    return quantity_estimate_data


def extract(excel_file):
    pattern = r"^\((\d+(\.\d+)?)\)$"
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = workbook.worksheets[0]
    sub_topics_dict = {}
    sub_topics_row, topics_row = None, None

    # Search for the "सि.नं." marker in the first two columns
    for row in sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=2):
        for cell in row:
            if cell.value in QuantityEstimateImportConstants.STARTING_COL_NAMES:
                topics_row = row
                break
        if topics_row:
            break
    if not topics_row:
        raise SNoNotFoundException(
            "Could not find the सि.नं. marker in the first two columns.\nMarkers: {}".format(
                QuantityEstimateImportConstants.STARTING_COL_NAMES
            )
        )

    start_index_row = topics_row[-1].row + 1

    quantity_estimate_data = []

    rows_to_process = list(sheet.iter_rows(min_row=start_index_row))

    columns_list = get_columns_list(
        QuantityEstimateImportConstants.OLD_QUANTITY_ESTIMATE_TOPICS_DICT
    )

    for row in rows_to_process:
        row_data = {"type": "data"}
        if row[0].value is None or str(row[0].value).strip() == "":
            total_column_value = row[columns_list.index("total")].value
            if total_column_value is not None:
                row_data["type"] = "total"
                try:
                    float(total_column_value)
                except:
                    continue
                last_s_no = quantity_estimate_data[-1]["s_no"]
                if "total" in last_s_no:
                    last_s_no = last_s_no.split(".")
                    last_s_no[-1] = str(int(last_s_no[-1]) + 1)
                    last_s_no = ".".join(last_s_no)
                else:
                    last_s_no = f"{last_s_no}.total.1"
                row_data["s_no"] = last_s_no
                total_column_value = float(f"{total_column_value:.5f}")
                row_data["total"] = total_column_value
                row_data["unit"] = row[columns_list.index("unit")].value
                quantity_estimate_data.append(row_data)
            continue
        for index, dict_items in enumerate(
            QuantityEstimateImportConstants.OLD_QUANTITY_ESTIMATE_TOPICS_DICT.items()
        ):
            dict_key, dict_value = dict_items
            if not isinstance(dict_value, dict):
                column_number = columns_list.index(dict_value)
                row_value = row[column_number].value
                row_data[dict_value] = english_nums(row_value)
                # row_data[dict_value] = row_value
            else:
                row_data[dict_value["code"]] = {}
                for sub_key_index, sub_key_items in enumerate(
                    dict_value["fields"].items()
                ):
                    sub_key, sub_value = sub_key_items
                    column_number = columns_list.index(sub_value)
                    value = row[column_number].value
                    if value is not None:
                        try:
                            float(value)
                        except:
                            match = re.match(pattern, str(value))
                            if match:
                                value = float(match.group(1))
                            else:
                                value = 0
                        value = float(f"{value:.5f}")
                    row_data[dict_value["code"]][sub_value] = value
        quantity_estimate_data.append(row_data)

    return quantity_estimate_data


def process(quantity_estimate_data):
    grouped_data = {}

    for item in quantity_estimate_data:
        s_no_parts = item["s_no"].split(".")
        current_level = grouped_data

        for part in s_no_parts:
            if part not in current_level:
                current_level[part] = {}
            current_level = current_level[part]

        current_level.update(item)

    return grouped_data


def import_old_template(excel_file):
    quantity_estimate_data = extract(excel_file)
    quantity_estimate_data = process(quantity_estimate_data)
    return quantity_estimate_data


def import_new_template(excel_file):
    quantity_estimate_data = extract_data_from_new_template(excel_file)
    quantity_estimate_data = process(quantity_estimate_data)
    return quantity_estimate_data


def import_quantity_estimate(excel_file):
    try:
        return import_new_template(excel_file)
    except:
        return import_old_template(excel_file)


def main():
    print("New Template")
    quantity_estimate_data = import_quantity_estimate("test_data/new imp-1 (1).xlsx")
    print(json.dumps(quantity_estimate_data, indent=4))
    print("-------------------------------------------")
    print("Old Template")
    quantity_estimate_data = import_quantity_estimate("test_data/new imp.xlsx")
    print(json.dumps(quantity_estimate_data, indent=4))


if __name__ == "__main__":
    main()
